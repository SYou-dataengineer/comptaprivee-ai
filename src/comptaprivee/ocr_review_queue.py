"""File locale des anomalies OCR à vérifier.

Cette file est séparée de la base SQLite des factures enregistrées :
elle contient uniquement des anomalies dérivées des exports PDF -> Excel/CSV.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from .ocr_table_validator import (
    detecter_anomalies_comptables,
    normaliser_montant_ocr,
)
from .review_queue import (
    ElementAVerifier,
    NiveauVerification,
)


CHEMIN_FILE_OCR_PAR_DEFAUT = (
    Path("data") / "ocr_review_queue.json"
)


@dataclass(frozen=True)
class FactureOCR:
    """Facture virtuelle issue d'une ligne OCR douteuse."""

    identifiant: int
    numero: str | None
    date: str | None
    fournisseur: str | None
    client: str | None
    sous_total: Decimal | None
    tps: Decimal | None
    tvq: Decimal | None
    total: Decimal | None
    date_creation: str
    source_document: str
    destination_export: str
    numero_page: int
    numero_tableau: int
    numero_ligne: int


def _normaliser_texte(valeur: object) -> str:
    return " ".join(
        str(valeur or "").strip().casefold().split()
    )


def _vers_decimal(valeur: object) -> Decimal | None:
    texte = normaliser_montant_ocr(
        str(valeur or "").strip()
    )

    if not texte:
        return None

    try:
        return Decimal(texte)
    except (InvalidOperation, ValueError):
        return None


def _charger(
    chemin_file: str | Path,
) -> list[dict]:
    chemin = Path(chemin_file)

    if not chemin.exists():
        return []

    try:
        contenu = json.loads(
            chemin.read_text(encoding="utf-8")
        )
    except (OSError, json.JSONDecodeError):
        # Fichier dérivé : une file invalide peut être reconstruite
        # lors de la prochaine conversion.
        return []

    if not isinstance(contenu, list):
        return []

    return [
        element
        for element in contenu
        if isinstance(element, dict)
    ]


def _enregistrer(
    elements: list[dict],
    chemin_file: str | Path,
) -> None:
    chemin = Path(chemin_file)
    chemin.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    temporaire = chemin.with_suffix(
        chemin.suffix + ".tmp"
    )

    temporaire.write_text(
        json.dumps(
            elements,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    temporaire.replace(chemin)


def _valeur_colonne(
    entete: list[str],
    ligne: list[str],
    *noms: str,
) -> str | None:
    entete_normalisee = [
        _normaliser_texte(v)
        for v in entete
    ]

    for nom in noms:
        nom_normalise = _normaliser_texte(nom)

        if nom_normalise not in entete_normalisee:
            continue

        index = entete_normalisee.index(
            nom_normalise
        )

        if index >= len(ligne):
            return None

        valeur = str(ligne[index] or "").strip()
        return valeur or None

    return None


def _identifiant(
    source: Path,
    numero_page: int,
    numero_tableau: int,
    numero_ligne: int,
) -> int:
    cle = (
        f"{source.resolve()}|{numero_page}|"
        f"{numero_tableau}|{numero_ligne}"
    )
    digest = hashlib.sha256(
        cle.encode("utf-8")
    ).hexdigest()

    # Les identifiants OCR sont négatifs afin de ne jamais entrer
    # en collision visuelle avec les identifiants SQLite.
    return -int(digest[:12], 16)


def _creer_alerte(
    *,
    source: Path,
    destination: Path,
    type_conversion: str,
    numero_page: int,
    numero_tableau: int,
    numero_ligne: int,
    entete: list[str],
    ligne: list[str],
) -> dict | None:
    entete_normalisee = [
        _normaliser_texte(v)
        for v in entete
    ]

    if "validation ocr" not in entete_normalisee:
        return None

    index_validation = entete_normalisee.index(
        "validation ocr"
    )

    if index_validation >= len(ligne):
        return None

    statut = _normaliser_texte(
        ligne[index_validation]
    )

    if statut not in {
        "à vérifier",
        "a vérifier",
        "a verifier",
    }:
        return None

    entete_donnees = [
        valeur
        for index, valeur in enumerate(entete)
        if index != index_validation
    ]
    ligne_donnees = [
        valeur
        for index, valeur in enumerate(ligne)
        if index != index_validation
    ]

    anomalies = detecter_anomalies_comptables(
        [
            entete_donnees,
            ligne_donnees,
        ]
    )

    raisons = [
        anomalie.message
        for anomalie in anomalies
    ]

    if not raisons:
        raisons.append(
            "Ligne marquée À VÉRIFIER par la validation OCR."
        )

    raisons.append(
        (
            f"Source OCR : {source.name}, "
            f"page {numero_page}, "
            f"tableau {numero_tableau}, "
            f"ligne {numero_ligne}."
        )
    )

    return {
        "identifiant": _identifiant(
            source,
            numero_page,
            numero_tableau,
            numero_ligne,
        ),
        "numero": _valeur_colonne(
            entete_donnees,
            ligne_donnees,
            "No facture",
            "Numéro de facture",
            "Numero de facture",
        ),
        "date": _valeur_colonne(
            entete_donnees,
            ligne_donnees,
            "Date",
        ),
        "fournisseur": _valeur_colonne(
            entete_donnees,
            ligne_donnees,
            "Fournisseur",
        ),
        "sous_total": _valeur_colonne(
            entete_donnees,
            ligne_donnees,
            "Sous-total",
            "Subtotal",
        ),
        "tps": _valeur_colonne(
            entete_donnees,
            ligne_donnees,
            "TPS",
        ),
        "tvq": _valeur_colonne(
            entete_donnees,
            ligne_donnees,
            "TVQ",
        ),
        "total": _valeur_colonne(
            entete_donnees,
            ligne_donnees,
            "Total",
        ),
        "source_document": str(source.resolve()),
        "destination_export": str(destination.resolve()),
        "type_conversion": type_conversion,
        "numero_page": numero_page,
        "numero_tableau": numero_tableau,
        "numero_ligne": numero_ligne,
        "raisons": raisons,
        "date_creation": datetime.now().isoformat(
            timespec="seconds"
        ),
    }


def _alertes_csv(
    source: Path,
    destination: Path,
    type_conversion: str,
) -> list[dict]:
    alertes: list[dict] = []

    with destination.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as fichier:
        lignes = list(csv.reader(fichier))

    numero_page = 1
    numero_tableau = 1
    entete: list[str] | None = None
    numero_ligne_tableau = 1

    for ligne in lignes:
        if not any(
            str(valeur).strip()
            for valeur in ligne
        ):
            entete = None
            numero_ligne_tableau = 1
            continue

        if (
            len(ligne) >= 2
            and str(ligne[0]).startswith("Page ")
            and str(ligne[1]).startswith("Tableau ")
        ):
            try:
                numero_page = int(
                    str(ligne[0]).split()[-1]
                )
                numero_tableau = int(
                    str(ligne[1]).split()[-1]
                )
            except ValueError:
                numero_page = 1
                numero_tableau = 1

            entete = None
            numero_ligne_tableau = 1
            continue

        normalisee = [
            _normaliser_texte(v)
            for v in ligne
        ]

        if "validation ocr" in normalisee:
            entete = [
                str(v or "")
                for v in ligne
            ]
            numero_ligne_tableau = 1
            continue

        if entete is None:
            continue

        numero_ligne_tableau += 1

        alerte = _creer_alerte(
            source=source,
            destination=destination,
            type_conversion=type_conversion,
            numero_page=numero_page,
            numero_tableau=numero_tableau,
            numero_ligne=numero_ligne_tableau,
            entete=entete,
            ligne=[
                str(v or "")
                for v in ligne
            ],
        )

        if alerte is not None:
            alertes.append(alerte)

    return alertes


def _alertes_excel(
    source: Path,
    destination: Path,
    type_conversion: str,
) -> list[dict]:
    from openpyxl import load_workbook

    alertes: list[dict] = []

    classeur = load_workbook(
        destination,
        data_only=True,
        read_only=True,
    )

    try:
        for feuille in classeur.worksheets:
            match = re.fullmatch(
                r"P(\d+)_T(\d+)",
                feuille.title,
            )

            if match:
                numero_page = int(match.group(1))
                numero_tableau = int(match.group(2))
            else:
                numero_page = 1
                numero_tableau = 1

            lignes = [
                [
                    "" if valeur is None else str(valeur)
                    for valeur in ligne
                ]
                for ligne in feuille.iter_rows(
                    values_only=True
                )
            ]

            if not lignes:
                continue

            entete = lignes[0]

            if "validation ocr" not in [
                _normaliser_texte(v)
                for v in entete
            ]:
                continue

            for numero_ligne, ligne in enumerate(
                lignes[1:],
                start=2,
            ):
                alerte = _creer_alerte(
                    source=source,
                    destination=destination,
                    type_conversion=type_conversion,
                    numero_page=numero_page,
                    numero_tableau=numero_tableau,
                    numero_ligne=numero_ligne,
                    entete=entete,
                    ligne=ligne,
                )

                if alerte is not None:
                    alertes.append(alerte)
    finally:
        classeur.close()

    return alertes


def synchroniser_export_ocr_a_verifier(
    source: str | Path,
    destination: str | Path,
    type_conversion: str,
    chemin_file: str | Path = CHEMIN_FILE_OCR_PAR_DEFAUT,
) -> int:
    """Synchronise un export OCR avec la file locale À vérifier."""
    source_path = Path(source)
    destination_path = Path(destination)

    if destination_path.suffix.lower() == ".csv":
        nouvelles = _alertes_csv(
            source_path,
            destination_path,
            type_conversion,
        )
    elif destination_path.suffix.lower() == ".xlsx":
        nouvelles = _alertes_excel(
            source_path,
            destination_path,
            type_conversion,
        )
    else:
        return 0

    source_resolue = str(
        source_path.resolve()
    )

    existantes = [
        element
        for element in _charger(chemin_file)
        if element.get("source_document")
        != source_resolue
    ]

    _enregistrer(
        [
            *existantes,
            *nouvelles,
        ],
        chemin_file,
    )

    return len(nouvelles)


def lister_alertes_ocr_a_verifier(
    chemin_file: str | Path = CHEMIN_FILE_OCR_PAR_DEFAUT,
) -> list[ElementAVerifier]:
    """Retourne les anomalies OCR sous la forme attendue par l'interface."""
    resultats: list[ElementAVerifier] = []

    for element in _charger(chemin_file):
        facture = FactureOCR(
            identifiant=int(
                element.get("identifiant") or 0
            ),
            numero=element.get("numero"),
            date=element.get("date"),
            fournisseur=element.get("fournisseur"),
            client=None,
            sous_total=_vers_decimal(
                element.get("sous_total")
            ),
            tps=_vers_decimal(
                element.get("tps")
            ),
            tvq=_vers_decimal(
                element.get("tvq")
            ),
            total=_vers_decimal(
                element.get("total")
            ),
            date_creation=str(
                element.get("date_creation") or ""
            ),
            source_document=str(
                element.get("source_document") or ""
            ),
            destination_export=str(
                element.get("destination_export") or ""
            ),
            numero_page=int(
                element.get("numero_page") or 1
            ),
            numero_tableau=int(
                element.get("numero_tableau") or 1
            ),
            numero_ligne=int(
                element.get("numero_ligne") or 1
            ),
        )

        raisons = tuple(
            str(raison)
            for raison in element.get(
                "raisons",
                [],
            )
        )

        resultats.append(
            ElementAVerifier(
                facture=facture,  # annotation informative seulement
                niveau=NiveauVerification.AVERTISSEMENT,
                raisons=raisons,
            )
        )

    return resultats

def marquer_alerte_ocr_resolue(
    identifiant: int,
    chemin_file: str | Path = CHEMIN_FILE_OCR_PAR_DEFAUT,
) -> bool:
    """Retire une anomalie OCR de la file locale après validation humaine."""
    elements = _charger(chemin_file)
    restants = [
        element
        for element in elements
        if int(element.get("identifiant") or 0) != identifiant
    ]

    if len(restants) == len(elements):
        return False

    _enregistrer(restants, chemin_file)
    return True
